#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import base64
import json
import mimetypes
import os
import platform
import re
import sqlite3
import subprocess
import sys
import threading
from copy import copy
from datetime import datetime
from http.server import BaseHTTPRequestHandler, HTTPServer
from io import BytesIO
from urllib.parse import quote, urlparse

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

try:
    from supabase import Client, create_client
except ImportError:
    Client = None
    create_client = None


TEMPLATE_PATH = os.environ.get("LEGACY_TEMPLATE_PATH", "")
DATA_ROOT = os.path.expanduser(os.environ.get("DATA_ROOT", "~/.tk_system"))
OUTPUT_DIR = os.path.expanduser(os.environ.get("OUTPUT_DIR", os.path.join(DATA_ROOT, "output")))
DB_PATH = os.path.join(DATA_ROOT, "tk_system.db")
TEMPLATES_DIR = os.path.join(DATA_ROOT, "templates")
SUPABASE_URL = (os.environ.get("SUPABASE_URL") or "").strip()
SUPABASE_SERVICE_ROLE_KEY = (os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_KEY") or "").strip()
SUPABASE_STATE_TABLE = os.environ.get("SUPABASE_STATE_TABLE", "tk_app_state")
SUPABASE_STATE_ROW_ID = os.environ.get("SUPABASE_STATE_ROW_ID", "primary")
SUPABASE_TEMPLATE_BUCKET = os.environ.get("SUPABASE_TEMPLATE_BUCKET", "tk-templates")
SUPABASE_ENABLED = bool(SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY)
ALLOWED_TEMPLATE_TYPES = {
    "estimate": "견적서",
    "delivery": "납품서",
    "invoice": "청구서",
    "statement": "거래명세서",
    "taxInvoice": "세금계산서",
}
DEFAULT_APP_DATA = {
    "customers": [],
    "items": [],
    "transactions": {},
    "payments": {},
}

_SUPABASE_CLIENT = None


def now_iso():
    return datetime.now().isoformat()


def normalize_app_data(data):
    normalized = dict(DEFAULT_APP_DATA)
    if isinstance(data, dict):
        for key in normalized:
            value = data.get(key)
            if isinstance(normalized[key], list):
                normalized[key] = value if isinstance(value, list) else []
            else:
                normalized[key] = value if isinstance(value, dict) else {}
    return normalized


def sanitize_filename(filename):
    base = os.path.basename(filename or "template.xlsx")
    return re.sub(r"[^A-Za-z0-9._-]", "_", base)


def sanitize_filename_token(value, fallback="-"):
    token = str(value or "").strip()
    if not token:
        return fallback
    # Keep Korean text but remove path-invalid characters.
    token = re.sub(r"[\\/:*?\"<>|]", "_", token)
    token = re.sub(r"\s+", " ", token).strip()
    return token or fallback


def parse_delivery_month(value):
    raw = str(value or "").strip()
    if not raw:
        return None

    for fmt in ("%Y-%m", "%Y-%m-%d", "%Y.%m", "%Y/%m"):
        try:
            parsed = datetime.strptime(raw, fmt)
            return datetime(parsed.year, parsed.month, 1)
        except ValueError:
            continue
    return None


def format_delivery_month(value):
    delivery_month = parse_delivery_month(value)
    if delivery_month:
        return f"{delivery_month.year}년 {delivery_month.month}월  일"

    raw = str(value or "").strip()
    if not raw:
        return ""
    return f"{raw}  일"


def number_to_korean(n):
    """정수를 한글 금액 표기로 변환 (예: 3000 → 삼천)"""
    n = int(round(float(n)))
    if n == 0:
        return '영'
    digits = ['', '일', '이', '삼', '사', '오', '육', '칠', '팔', '구']
    units = ['', '십', '백', '천']
    big_units = ['', '만', '억', '조']
    result = ''
    big_unit_idx = 0
    while n > 0:
        chunk = n % 10000
        if chunk != 0:
            chunk_str = ''
            for pos in range(3, -1, -1):
                digit = (chunk // (10 ** pos)) % 10
                if digit == 0:
                    continue
                if digit == 1 and pos > 0:
                    chunk_str += units[pos]
                else:
                    chunk_str += digits[digit] + units[pos]
            result = chunk_str + big_units[big_unit_idx] + result
        n //= 10000
        big_unit_idx += 1
    return result


def format_total_amount(total):
    amount = int(round(float(total)))
    return f"₩ {amount:,} (金 {number_to_korean(amount)}원정)"


def apply_estimate_overrides(ws, data):
    total = sum(
        float(item.get("quantity", 0) or 0) * float(item.get("unitPrice", 0) or 0)
        for item in data.get("items", [])
    )

    delivery_month_text = format_delivery_month(data.get("deliveryMonth", ""))
    if delivery_month_text:
        ws["A2"].value = delivery_month_text

    ws["A7"].value = total
    ws["A7"].number_format = '"합계금액:"#,##0"원"'


def build_estimate_filename(data):
    return "estimate.xlsx"


def ensure_local_dirs():
    os.makedirs(DATA_ROOT, exist_ok=True)
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def copy_cell_style(source_cell, target_cell):
    try:
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.number_format:
            target_cell.number_format = copy(source_cell.number_format)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
    except Exception:
        pass


def set_cell_value_safe(ws, row_num, col_idx, value):
    cell = ws.cell(row=row_num, column=col_idx)
    if isinstance(cell, MergedCell):
        for merged in ws.merged_cells.ranges:
            if (
                merged.min_row <= row_num <= merged.max_row
                and merged.min_col <= col_idx <= merged.max_col
            ):
                ws.cell(row=merged.min_row, column=merged.min_col).value = value
                return
        return
    cell.value = value


def replace_variables(ws, data):
    items = data.get("items", [])

    # 품목 템플릿 행 자동 감지 ({품목명} 플레이스홀더가 있는 행)
    item_template_row = None
    item_col_map = {}  # placeholder → column index

    ITEM_PLACEHOLDERS = {
        "{품목명}": "product",
        "{규격}": "spec",
        "{수량}": "quantity",
        "{단위}": "unit",
        "{단가}": "unitPrice",
        "{금액}": "amount",
        "{공급가액}": "amount",
        "{비고}": "remark",
    }

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "{품목명}" in cell.value:
                item_template_row = cell.row
                break
        if item_template_row:
            break

    # 품목 행 컬럼 매핑 구성
    if item_template_row:
        for cell in ws[item_template_row]:
            if cell.value and isinstance(cell.value, str):
                for ph, field in ITEM_PLACEHOLDERS.items():
                    if ph in cell.value:
                        item_col_map[cell.column] = field

    # 품목 데이터 채우기 (품목 개수에 맞춰 행을 늘려 하단 영역을 자동으로 아래로 이동)
    if item_template_row and item_col_map:
        render_item_count = max(len(items), 1)
        insert_count = max(render_item_count - 1, 0)

        if insert_count > 0:
            ws.insert_rows(item_template_row + 1, amount=insert_count)

            template_row_height = ws.row_dimensions[item_template_row].height
            template_row_merges = [
                merged for merged in ws.merged_cells.ranges
                if merged.min_row == item_template_row and merged.max_row == item_template_row
            ]

            for offset in range(1, render_item_count):
                row_num = item_template_row + offset
                if template_row_height is not None:
                    ws.row_dimensions[row_num].height = template_row_height

                for col_idx in range(1, ws.max_column + 1):
                    source_cell = ws.cell(row=item_template_row, column=col_idx)
                    target_cell = ws.cell(row=row_num, column=col_idx)
                    copy_cell_style(source_cell, target_cell)

                for merged in template_row_merges:
                    ws.merge_cells(
                        start_row=row_num,
                        start_column=merged.min_col,
                        end_row=row_num,
                        end_column=merged.max_col,
                    )

        for idx in range(render_item_count):
            row_num = item_template_row + idx
            item = items[idx] if idx < len(items) else None

            for col_idx, field in item_col_map.items():
                if item:
                    if field == "quantity":
                        set_cell_value_safe(ws, row_num, col_idx, float(item.get("quantity", 0) or 0))
                    elif field == "unitPrice":
                        set_cell_value_safe(ws, row_num, col_idx, float(item.get("unitPrice", 0) or 0))
                    elif field == "amount":
                        qty = float(item.get("quantity", 0) or 0)
                        price = float(item.get("unitPrice", 0) or 0)
                        set_cell_value_safe(ws, row_num, col_idx, qty * price)
                    else:
                        set_cell_value_safe(ws, row_num, col_idx, item.get(field, ""))
                else:
                    set_cell_value_safe(ws, row_num, col_idx, "")

    # 합계 계산
    total = sum(
        float(item.get("quantity", 0) or 0) * float(item.get("unitPrice", 0) or 0)
        for item in items
    )

    # 나머지 플레이스홀더 치환 (품목 행 제외)
    customer = data.get("customer", "")
    replacements = {
        "{납품월}": format_delivery_month(data.get("deliveryMonth", "")),
        "{거래처명}": customer,
        "{공사명}": data.get("projectName", ""),
        "{합계}": format_total_amount(total),
    }

    # 마지막 {거래처명} 셀 위치 미리 탐색
    last_customer_cell = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "{거래처명}" in cell.value:
                is_item_ph = any(ph in cell.value for ph in ITEM_PLACEHOLDERS)
                if not is_item_ph:
                    last_customer_cell = (cell.row, cell.column)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                # 품목 플레이스홀더가 남은 셀은 빈 문자열로 처리
                is_item_placeholder = any(ph in cell.value for ph in ITEM_PLACEHOLDERS)
                if is_item_placeholder:
                    set_cell_value_safe(ws, cell.row, cell.column, "")
                    continue
                # 마지막 {거래처명}: 각 글자를 공백 4칸으로 분리
                if "{거래처명}" in cell.value and last_customer_cell == (cell.row, cell.column):
                    ph = "{거래처명}"
                    idx = cell.value.find(ph)
                    prefix = cell.value[:idx].rstrip()
                    suffix = cell.value[idx + len(ph):].lstrip()
                    all_chars = list(prefix) + list(customer) + list(suffix)
                    spaced = '    '.join(all_chars)
                    set_cell_value_safe(ws, cell.row, cell.column, spaced)
                    continue
                for placeholder, value in replacements.items():
                    if placeholder in cell.value:
                        set_cell_value_safe(ws, cell.row, cell.column, cell.value.replace(placeholder, str(value)))
    
    # 모든 셀값 정규화 (한글 문자열이 올바르게 처리되도록)
    import unicodedata
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                try:
                    # NFC 정규형으로 정규화
                    cell.value = unicodedata.normalize('NFC', cell.value)
                except Exception:
                    pass


def get_supabase_client():
    global _SUPABASE_CLIENT

    if not SUPABASE_ENABLED:
        return None

    if create_client is None:
        raise RuntimeError("supabase 패키지가 설치되지 않았습니다. requirements.txt를 다시 설치해주세요.")

    if _SUPABASE_CLIENT is None:
        _SUPABASE_CLIENT = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
    return _SUPABASE_CLIENT


def ensure_supabase_bucket():
    client = get_supabase_client()
    if not client:
        return

    try:
        client.storage.get_bucket(SUPABASE_TEMPLATE_BUCKET)
    except Exception:
        client.storage.create_bucket(
            SUPABASE_TEMPLATE_BUCKET,
            options={
                "public": False,
                "allowed_mime_types": [
                    "application/vnd.ms-excel",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/vnd.ms-excel.sheet.macroEnabled.12",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.template",
                    "application/vnd.ms-excel.template.macroEnabled.12",
                ],
                "file_size_limit": 20 * 1024 * 1024,
            },
        )


def verify_supabase_table():
    client = get_supabase_client()
    if not client:
        return

    try:
        client.table(SUPABASE_STATE_TABLE).select("id").limit(1).execute()
    except Exception as error:
        raise RuntimeError(
            "Supabase 테이블이 준비되지 않았습니다. README의 supabase_schema.sql을 먼저 실행해주세요. "
            f"원인: {error}"
        ) from error


def get_local_template_path(template_type):
    ensure_local_dirs()
    prefix = f"{template_type}__"
    candidates = []
    for name in os.listdir(TEMPLATES_DIR):
        if name.startswith(prefix):
            full_path = os.path.join(TEMPLATES_DIR, name)
            if os.path.isfile(full_path):
                candidates.append(full_path)

    if not candidates:
        return None

    candidates.sort(key=os.path.getmtime, reverse=True)
    return candidates[0]


def get_supabase_template_entries(template_type):
    client = get_supabase_client()
    if not client:
        return []

    entries = client.storage.from_(SUPABASE_TEMPLATE_BUCKET).list(
        template_type,
        {
            "limit": 100,
            "offset": 0,
            "sortBy": {"column": "name", "order": "desc"},
        },
    ) or []

    normalized = []
    for entry in entries:
        file_name = entry.get("name")
        if not file_name:
            continue
        metadata = entry.get("metadata") or {}
        normalized.append(
            {
                "objectPath": f"{template_type}/{file_name}",
                "fileName": file_name,
                "savedAt": entry.get("updated_at") or entry.get("created_at"),
                "size": metadata.get("size") or entry.get("size") or 0,
            }
        )

    normalized.sort(key=lambda item: item.get("savedAt") or "", reverse=True)
    return normalized


def get_template_info(template_type):
    if SUPABASE_ENABLED:
        entries = get_supabase_template_entries(template_type)
        if entries:
            return entries[0]

    path = get_local_template_path(template_type)
    if path and os.path.exists(path):
        stat = os.stat(path)
        return {
            "objectPath": path,
            "fileName": os.path.basename(path).split("__", 1)[-1],
            "savedAt": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            "size": stat.st_size,
        }

    return None


def read_template_bytes(template_type):
    info = get_template_info(template_type)
    if info:
        if SUPABASE_ENABLED and info["objectPath"].startswith(f"{template_type}/"):
            content = get_supabase_client().storage.from_(SUPABASE_TEMPLATE_BUCKET).download(info["objectPath"])
            return info["fileName"], content

        with open(info["objectPath"], "rb") as file_obj:
            return info["fileName"], file_obj.read()

    if template_type == "estimate" and TEMPLATE_PATH and os.path.exists(TEMPLATE_PATH):
        with open(TEMPLATE_PATH, "rb") as file_obj:
            return os.path.basename(TEMPLATE_PATH), file_obj.read()

    return None, None


def list_templates():
    result = {}
    for template_type, label in ALLOWED_TEMPLATE_TYPES.items():
        info = get_template_info(template_type)
        if info:
            result[template_type] = {
                "label": label,
                "uploaded": True,
                "fileName": info["fileName"],
                "savedAt": info["savedAt"],
                "size": info["size"],
            }
        else:
            result[template_type] = {
                "label": label,
                "uploaded": False,
                "fileName": None,
                "savedAt": None,
                "size": 0,
            }
    return result


def save_template_file_local(template_type, file_name, content):
    ensure_local_dirs()
    safe_name = sanitize_filename(file_name)
    existing = get_local_template_path(template_type)
    if existing and os.path.exists(existing):
        os.remove(existing)

    target_path = os.path.join(TEMPLATES_DIR, f"{template_type}__{safe_name}")
    with open(target_path, "wb") as file_obj:
        file_obj.write(content)
    return target_path


def save_template_file_supabase(template_type, file_name, content):
    client = get_supabase_client()
    safe_name = sanitize_filename(file_name)
    current_entries = get_supabase_template_entries(template_type)
    if current_entries:
        client.storage.from_(SUPABASE_TEMPLATE_BUCKET).remove([entry["objectPath"] for entry in current_entries])

    object_path = f"{template_type}/{safe_name}"
    mime_type = mimetypes.guess_type(safe_name)[0] or "application/octet-stream"
    client.storage.from_(SUPABASE_TEMPLATE_BUCKET).upload(
        path=object_path,
        file=content,
        file_options={
            "content-type": mime_type,
            "upsert": "true",
        },
    )
    return object_path


def save_template_file(template_type, file_name, content_base64):
    if template_type not in ALLOWED_TEMPLATE_TYPES:
        raise ValueError("invalid template type")

    content = base64.b64decode(content_base64.encode("utf-8"), validate=True)
    if SUPABASE_ENABLED:
        return save_template_file_supabase(template_type, file_name, content)
    return save_template_file_local(template_type, file_name, content)


def delete_template_file(template_type):
    deleted = False

    if SUPABASE_ENABLED:
        entries = get_supabase_template_entries(template_type)
        if entries:
            get_supabase_client().storage.from_(SUPABASE_TEMPLATE_BUCKET).remove([entry["objectPath"] for entry in entries])
            deleted = True

    path = get_local_template_path(template_type)
    if path and os.path.exists(path):
        os.remove(path)
        deleted = True

    return deleted


def init_db_local():
    ensure_local_dirs()
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS app_snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                saved_at TEXT NOT NULL,
                data_json TEXT NOT NULL
            )
            """
        )
        conn.commit()
    finally:
        conn.close()


def save_app_data_local(data):
    payload = json.dumps(normalize_app_data(data), ensure_ascii=False)
    saved_at = now_iso()
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute(
            "INSERT INTO app_snapshots (saved_at, data_json) VALUES (?, ?)",
            (saved_at, payload),
        )
        conn.commit()
    finally:
        conn.close()
    return saved_at


def load_latest_app_data_local():
    if not os.path.exists(DB_PATH):
        return {"data": dict(DEFAULT_APP_DATA), "savedAt": None}

    conn = sqlite3.connect(DB_PATH)
    try:
        row = conn.execute(
            "SELECT data_json, saved_at FROM app_snapshots ORDER BY id DESC LIMIT 1"
        ).fetchone()
    finally:
        conn.close()

    if not row:
        return {"data": dict(DEFAULT_APP_DATA), "savedAt": None}

    try:
        data = json.loads(row[0]) if row[0] else {}
    except Exception:
        data = {}

    return {"data": normalize_app_data(data), "savedAt": row[1]}


def save_app_data_supabase(data):
    client = get_supabase_client()
    saved_at = now_iso()
    payload = {
        "id": SUPABASE_STATE_ROW_ID,
        "data_json": normalize_app_data(data),
        "updated_at": saved_at,
    }
    client.table(SUPABASE_STATE_TABLE).upsert(payload, on_conflict="id").execute()
    return saved_at


def load_latest_app_data_supabase():
    client = get_supabase_client()
    response = (
        client.table(SUPABASE_STATE_TABLE)
        .select("id, data_json, updated_at")
        .eq("id", SUPABASE_STATE_ROW_ID)
        .maybe_single()
        .execute()
    )

    row = response.data or None
    if not row:
        return {"data": dict(DEFAULT_APP_DATA), "savedAt": None}

    data = row.get("data_json")
    if isinstance(data, str):
        try:
            data = json.loads(data)
        except Exception:
            data = {}

    return {
        "data": normalize_app_data(data),
        "savedAt": row.get("updated_at"),
    }


def save_app_data(data):
    if SUPABASE_ENABLED:
        return save_app_data_supabase(data)
    return save_app_data_local(data)


def load_latest_app_data():
    if SUPABASE_ENABLED:
        return load_latest_app_data_supabase()
    return load_latest_app_data_local()


def seed_supabase_from_local():
    if not SUPABASE_ENABLED:
        return

    remote_state = load_latest_app_data_supabase()
    local_state = load_latest_app_data_local()
    if not remote_state["savedAt"] and local_state["savedAt"]:
        save_app_data_supabase(local_state["data"])

    for template_type in ALLOWED_TEMPLATE_TYPES:
        remote_info = get_template_info(template_type)
        local_path = get_local_template_path(template_type)
        if remote_info and remote_info["objectPath"].startswith(f"{template_type}/"):
            continue
        if local_path and os.path.exists(local_path):
            with open(local_path, "rb") as file_obj:
                save_template_file_supabase(
                    template_type,
                    os.path.basename(local_path).split("__", 1)[-1],
                    file_obj.read(),
                )


def generate_estimate(data):
    return generate_document(data, "estimate")


def generate_document(data, doc_type):
    try:
        if doc_type not in ALLOWED_TEMPLATE_TYPES:
            raise ValueError("지원하지 않는 문서 타입입니다.")

        template_name, template_bytes = read_template_bytes(doc_type)
        if not template_bytes:
            raise FileNotFoundError(
                f"{ALLOWED_TEMPLATE_TYPES[doc_type]} 템플릿이 없습니다. 서식업로드에서 먼저 업로드해주세요."
            )

        wb = load_workbook(BytesIO(template_bytes))
        ws = wb.active
        replace_variables(ws, data)

        if doc_type == "estimate":
            apply_estimate_overrides(ws, data)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        ensure_local_dirs()
        if doc_type == "estimate":
            filename = build_estimate_filename(data)
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            customer_name = sanitize_filename_token(data.get("customer"), "document")
            filename = f"{customer_name}_{doc_type}_{timestamp}.xlsx"

        filepath = None
        # Optional local save for troubleshooting/manual archive.
        if os.environ.get("SAVE_GENERATED_LOCAL") == "1":
            filepath = os.path.join(OUTPUT_DIR, filename)
            with open(filepath, "wb") as file_obj:
                file_obj.write(output.getvalue())

        return output.getvalue(), filename, filepath
    except Exception as error:
        print(f"Error: {error}")
        import traceback

        traceback.print_exc()
        return None, None, None


def open_file(filepath):
    try:
        if os.environ.get("RENDER") or os.environ.get("DISABLE_AUTO_OPEN") == "1":
            return
        if platform.system() == "Darwin":
            subprocess.Popen(["open", filepath], encoding='utf-8')
        elif platform.system() == "Windows":
            os.startfile(filepath)
        elif platform.system() == "Linux":
            subprocess.Popen(["xdg-open", filepath], encoding='utf-8')
    except Exception as error:
        print(f"Warning: {error}")


class EstimateHandler(BaseHTTPRequestHandler):
    def _set_file_download_headers(self, filename):
        safe_filename = sanitize_filename(filename or "document.xlsx")
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{safe_filename}"')
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Expose-Headers", "Content-Disposition")

    def _serve_static_file(self, relative_path):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        safe_relative = os.path.normpath(relative_path).lstrip("/\\")
        file_path = os.path.abspath(os.path.join(base_dir, safe_relative))

        if not file_path.startswith(base_dir + os.sep):
            self.send_response(403)
            self.end_headers()
            return True

        if not os.path.isfile(file_path):
            return False

        content_type, _ = mimetypes.guess_type(file_path)
        if not content_type:
            content_type = "application/octet-stream"

        with open(file_path, "rb") as file_obj:
            payload = file_obj.read()

        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)
        return True

    def _send_json(self, status_code, body):
        self.send_response(status_code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, DELETE, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
        self.wfile.write(json.dumps(body, ensure_ascii=False).encode("utf-8"))

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, DELETE, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self):
        try:
            request_path = urlparse(self.path).path

            if request_path == "/api/templates":
                self._send_json(200, {"templates": list_templates()})
                return

            if request_path.startswith("/api/templates/") and request_path.endswith("/download"):
                parts = request_path.split("/")
                if len(parts) >= 5:
                    template_type = parts[3]
                    if template_type not in ALLOWED_TEMPLATE_TYPES:
                        self._send_json(404, {"error": "template type not found"})
                        return

                    filename, file_bytes = read_template_bytes(template_type)
                    if not file_bytes:
                        self._send_json(404, {"error": "template file not found"})
                        return

                    self.send_response(200)
                    self.send_header("Content-Type", "application/octet-stream")
                    self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                    self.send_header("Access-Control-Allow-Origin", "*")
                    self.end_headers()
                    self.wfile.write(file_bytes)
                    return

            if request_path == "/api/data":
                self._send_json(200, load_latest_app_data())
                return

            if request_path == "/api/health":
                self._send_json(
                    200,
                    {
                        "ok": True,
                        "storage": "supabase" if SUPABASE_ENABLED else "local",
                        "supabaseConfigured": SUPABASE_ENABLED,
                    },
                )
                return

            if request_path in ["/", "/index.html"]:
                if self._serve_static_file("index.html"):
                    return

            if request_path.startswith("/css/") or request_path.startswith("/js/") or request_path.startswith("/pages/"):
                if self._serve_static_file(request_path.lstrip("/")):
                    return

            self.send_response(404)
            self.end_headers()
        except Exception as error:
            self._send_json(500, {"ok": False, "error": str(error)})

    def do_POST(self):
        if self.path == "/api/templates":
            content_length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(content_length)
            try:
                payload = json.loads(body.decode("utf-8"))
                template_type = payload.get("templateType")
                file_name = payload.get("fileName")
                content_base64 = payload.get("contentBase64")

                if not template_type or not file_name or not content_base64:
                    raise ValueError("templateType, fileName, contentBase64 are required")

                save_template_file(template_type, file_name, content_base64)
                self._send_json(200, {"ok": True, "templates": list_templates()})
            except Exception as error:
                self._send_json(400, {"ok": False, "error": str(error)})
            return

        if self.path == "/api/data":
            content_length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(content_length)

            try:
                data = json.loads(body.decode("utf-8"))
                if not isinstance(data, dict):
                    raise ValueError("data must be an object")

                saved_at = save_app_data(data)
                self._send_json(200, {"ok": True, "savedAt": saved_at})
            except Exception as error:
                self._send_json(400, {"ok": False, "error": str(error)})
            return

        if self.path == "/estimate":
            content_length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(content_length)

            try:
                data = json.loads(body.decode("utf-8"))
                file_data, filename, filepath = generate_estimate(data)

                if file_data:
                    if filepath:
                        threading.Thread(target=open_file, args=(filepath,), daemon=True).start()
                    self.send_response(200)
                    self._set_file_download_headers(filename)
                    self.end_headers()
                    self.wfile.write(file_data)
                else:
                    self._send_json(500, {"error": "파일 생성 실패"})
            except Exception as error:
                print(f"Error: {error}")
                self._send_json(500, {"error": str(error)})
            return

        if self.path == "/document":
            content_length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(content_length)

            try:
                payload = json.loads(body.decode("utf-8"))
                doc_type = payload.get("docType")
                data = payload.get("data")

                if not doc_type or not isinstance(data, dict):
                    raise ValueError("docType, data are required")

                file_data, filename, filepath = generate_document(data, doc_type)

                if file_data:
                    if filepath:
                        threading.Thread(target=open_file, args=(filepath,), daemon=True).start()
                    self.send_response(200)
                    self._set_file_download_headers(filename)
                    self.end_headers()
                    self.wfile.write(file_data)
                else:
                    self._send_json(500, {"error": "파일 생성 실패"})
            except Exception as error:
                print(f"Error: {error}")
                self._send_json(500, {"error": str(error)})
            return

        self.send_response(404)
        self.end_headers()

    def do_DELETE(self):
        if self.path.startswith("/api/templates/"):
            parts = self.path.split("/")
            if len(parts) >= 4:
                template_type = parts[3]
                if template_type not in ALLOWED_TEMPLATE_TYPES:
                    self._send_json(404, {"ok": False, "error": "template type not found"})
                    return
                try:
                    deleted = delete_template_file(template_type)
                    self._send_json(200, {"ok": True, "deleted": deleted, "templates": list_templates()})
                except Exception as error:
                    self._send_json(400, {"ok": False, "error": str(error)})
                return

        self.send_response(404)
        self.end_headers()

    def log_message(self, format, *args):
        pass


def initialize_storage():
    ensure_local_dirs()
    init_db_local()

    if SUPABASE_ENABLED:
        ensure_supabase_bucket()
        verify_supabase_table()
        seed_supabase_from_local()


def run_server(port=5050, host="0.0.0.0"):
    initialize_storage()
    server = HTTPServer((host, port), EstimateHandler)
    storage_label = "Supabase" if SUPABASE_ENABLED else "Local"
    print(f"견적서 서버 시작: http://{host}:{port} ({storage_label})")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("서버 종료")
        server.shutdown()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5050"))
    if len(sys.argv) > 1:
        try:
            port = int(sys.argv[1])
        except ValueError:
            print("포트는 숫자여야 합니다. 기본값 5050을 사용합니다.")
            port = 5050
    run_server(port)
