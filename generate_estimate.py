#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import sys
import os
from openpyxl import load_workbook
from copy import copy
from datetime import datetime
import subprocess
import platform

# 템플릿 파일 경로
TEMPLATE_PATH = "/Users/stark1313/시스템개발/01. 사무실 주문관리 시스템/서식업로드/양식_견적서.xlsx"
OUTPUT_DIR = os.path.expanduser("~/Desktop")  # 바탕화면에 저장

def copy_cell_style(source_cell, target_cell):
    """셀 스타일 복사"""
    try:
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.number_format:
            target_cell.number_format = copy(source_cell.number_format)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
    except:
        pass

def replace_variables(ws, data):
    """워크시트의 변수 치환"""
    # 단순 변수 치환: {변수명}
    replacements = {
        '{납품월}': data.get('deliveryMonth', ''),
        '{거래처명}': data.get('customer', ''),
        '{공사명}': data.get('projectName', ''),
    }
    
    # 셀 값 중에 변수가 있으면 치환
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for placeholder, value in replacements.items():
                    if placeholder in cell.value:
                        cell.value = cell.value.replace(placeholder, str(value))
    
    # 품목 데이터 입력 (행 9부터 시작)
    items = data.get('items', [])
    start_row = 9  # 첫 데이터 행
    template_row = 9  # 템플릿 스타일 참고 행
    
    for idx, item in enumerate(items[:20]):  # 최대 20개 품목
        row_num = start_row + idx
        
        # 템플릿 행에서 스타일 복사
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            source_cell = ws[f'{col_letter}{template_row}']
            target_cell = ws[f'{col_letter}{row_num}']
            copy_cell_style(source_cell, target_cell)
        
        # 데이터 입력
        ws[f'A{row_num}'].value = item.get('product', '')
        ws[f'B{row_num}'].value = item.get('spec', '')
        ws[f'C{row_num}'].value = float(item.get('quantity', 0) or 0)
        ws[f'D{row_num}'].value = item.get('unit', '')
        ws[f'E{row_num}'].value = float(item.get('unitPrice', 0) or 0)
        
        # 금액 = 수량 * 단가
        quantity = float(item.get('quantity', 0) or 0)
        unit_price = float(item.get('unitPrice', 0) or 0)
        amount = quantity * unit_price
        ws[f'F{row_num}'].value = amount
        ws[f'G{row_num}'].value = item.get('remark', '')

def generate_estimate(data):
    """견적서 Excel 생성"""
    try:
        # 템플릿 로드
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
        
        # 변수 치환
        replace_variables(ws, data)
        
        # 합계 계산
        total = sum(float(ws[f'F{row}'].value or 0) for row in range(9, 29))
        
        # 합계 셀에 입력 (기존 포뮬라가 있는 곳)
        ws['F11'].value = total
        
        # 파일 저장
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        customer_name = data.get('customer', 'estimate').replace('/', '_')
        filename = f"{customer_name}_견적서_{timestamp}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        wb.save(filepath)
        
        print(f"SUCCESS:{filepath}")
        return filepath
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        return None

def open_file(filepath):
    """파일 자동 실행"""
    try:
        if platform.system() == 'Darwin':  # macOS
            subprocess.Popen(['open', filepath])
        elif platform.system() == 'Windows':
            os.startfile(filepath)
        elif platform.system() == 'Linux':
            subprocess.Popen(['xdg-open', filepath])
    except Exception as e:
        print(f"WARNING: 파일 열기 실패: {e}")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 generate_estimate.py '<json_data>'")
        sys.exit(1)
    
    try:
        # JSON 데이터 파싱
        json_data = sys.argv[1]
        data = json.loads(json_data)
        
        # 견적서 생성
        filepath = generate_estimate(data)
        
        if filepath:
            # 파일 자동 실행
            open_file(filepath)
        else:
            print("ERROR: 파일 생성 실패")
            sys.exit(1)
            
    except json.JSONDecodeError as e:
        print(f"ERROR: JSON 파싱 실패: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)
